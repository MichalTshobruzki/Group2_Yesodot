import xlrd
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
import numpy as np
#ITSEMIL

'''this func recieves the meassage from the shift manager'''
def MessageForManager(access):
    messages_list = []
    row_list = []
    message_loc = r'C:\Users\micha\Desktop\קוד מיכל\Group2_Yesodot\workOnExcel\messages.xlsx'
    message_file = xlrd.open_workbook(message_loc)
    sheet = message_file.sheet_by_index(0)
    for i in range(0, sheet.nrows):
        row_list = sheet.row_values(i)
        if i > 0:
            row_list[0] = int(row_list[0])
        messages_list.append(row_list)
    print("Enter here you message: ")
    messageFromWorker = input()
    messages_list.append([sheet.nrows, messageFromWorker])

    message_workbook = xlsxwriter.Workbook('messages.xlsx')
    worksheet = message_workbook.add_worksheet('messages')

    for i in range(len(messages_list)):
        worksheet.write(i, 0, messages_list[i][0])
        worksheet.write(i, 1, messages_list[i][1])
    message_workbook.close()

    Open_Menu(access)



'''find a custumer in the members club'''
def find_custumer(access):
    name, last = input('enter the first name: '), input('enter the last name: ')
    file_loc = r'C:\Users\micha\Desktop\קוד מיכל\Group2_Yesodot\Hackathon\membership.xlsx'
    workbook = xlrd.open_workbook(file_loc)
    worksheet = workbook.sheet_by_index(0)
    worksheet.cell_value(0, 0)
    for i in range(worksheet.nrows):
        if worksheet.cell_value(i, 0) == name and worksheet.cell_value(i, 1) == last:
            return('Exist!')
    return('Doesnt Exist!')



'''add worker Constraints'''
def add_worker_Constraints(access):
    name = input('Enter your name: ')
    workbook = xlsxwriter.Workbook('Constraints.xlsx')
    worksheet = workbook.add_worksheet(name)

    worksheet.write(1, 0, 'Morning')
    worksheet.write(2, 0, 'Evening')
    worksheet.write(0, 1, 'Sunday')
    worksheet.write(0, 2, 'Monday')
    worksheet.write(0, 3, 'Tuesday')
    worksheet.write(0, 4, 'Wednesday')
    worksheet.write(0, 5, 'Thursday')
    worksheet.write(0, 6, 'Friday')
    worksheet.write(0, 7, 'Saturday')

    constraint1_day, constraint1_shift = input('enter your first constraint-> day: '), \
                                         input('enter your first constraint-> shift: ')
    constraint2_day, constraint2_shift = input('enter your second constraint-> day: '), \
                                         input('enter your second constraint-> shift: ')

    if constraint1_day == 'Sunday':
        constraint1_day = 1
    if constraint1_day == 'Monday':
        constraint1_day = 2
    if constraint1_day == 'Tuesday':
        constraint1_day = 3
    if constraint1_day == 'Wednesday':
        constraint1_day = 4
    if constraint1_day == 'Thursday':
        constraint1_day = 5
    if constraint1_day == 'Friday':
        constraint1_day = 6
    if constraint1_day == 'Saturday':
        constraint1_day = 7
    if constraint1_shift == 'Morning':
        constraint1_shift = 1
    if constraint1_shift == 'Evening':
        constraint1_shift = 2

    if constraint2_day == 'Sunday':
        constraint2_day = 1
    if constraint2_day == 'Monday':
        constraint2_day = 2
    if constraint2_day == 'Tuesday':
        constraint2_day = 3
    if constraint2_day == 'Wednesday':
        constraint2_day = 4
    if constraint2_day == 'Thursday':
        constraint2_day = 5
    if constraint2_day == 'Friday':
        constraint2_day = 6
    if constraint2_day == 'Saturday':
        constraint2_day = 7
    if constraint2_shift == 'Morning':
        constraint2_shift = 1
    if constraint2_shift == 'Evening':
        constraint2_shift = 2

    worksheet.write(constraint1_shift, constraint1_day, 'NO')
    worksheet.write(constraint2_shift, constraint2_day, 'NO')
    workbook.close()
    Open_Menu(access)


def Open_Menu(access):
    access_manage = 'manager'
    access_Responsible = 'shift r'
    access_worker ='worker'


    if access == access_manage:
        manager_menu(access)
    if access == access_Responsible:
        Responsible_menu(access)
    if access == access_worker:
        worker_menu(access)

def manager_menu(access):
    file_loc = r'C:\Users\micha\Desktop\קוד מיכל\Group2_Yesodot\workOnExcel\messages.xlsx'
    workbook = xlrd.open_workbook(file_loc)
    worksheet = workbook.sheet_by_index(0)
    print('-----------------------------------------------')
    print("*****Dear manager,you have new alert*****")
    print('*****{0}*****'.format(worksheet.cell_value(worksheet.nrows-1, 1)))
    print('manager menu:')
    print('Select the desired action ')
    print('1- sell item')
    print('2- Issue reports')
    print('3- Cancelling a transaction\ Refund')
    print('4- Replenishment')
    print('5- Remove item inventory')
    print('6- Changes in work arrangements')
    print('7- add customer to customer club')
    print('8- remove customer from customer club')
    print('-----------------------------------------------')

def Responsible_menu(access):
    print('-----------------------------------------------')
    print('responsible menu:')
    print('Select the desired action ')
    print('1- sell item')
    print('2- Issue reports')
    print('3- Submit messages to the administrator')
    print('4- Submission of constraints')
    print('5- add customer to customer club')
    print('6- remove customer from customer club')
    print('-----------------------------------------------')

    choice = input('your choice: ')
    if choice == '4':
        print('1- submission of constrains')
        print('2- Viewing constraints')
        print('-----------------------------------------------')
        choice = input()
        if choice == '1':
            add_worker_Constraints(access)
    if choice == '3':
        MessageForManager(access)


def worker_menu(access):
    print('-----------------------------------------------')
    print('worker menu:')
    print('Select the desired action ')
    print('1- sell item')
    print('2- Issue reports')
    print('3- Closing the POS')
    print('4- Submission of constraints')
    print('5- add customer to customer club')
    print('6- find customer in customer club')
    print('-----------------------------------------------')
    choice = input()
    if choice == '4':
        print('1- submission of constrains')
        print('2- Viewing constraints')
        print('-----------------------------------------------')
        choice = input()
        if choice == '1':
            add_worker_Constraints(access)
    if choice == '6':
        print(find_custumer(access))
        Open_Menu(access)


def Error_page():
    exit(0)

def Log_In():
    file_loc = r'C:\Users\micha\Desktop\קוד מיכל\Group2_Yesodot\Hackathon\passwarde.xlsx'

    pas_file = xlrd.open_workbook(file_loc)
    sheet = pas_file.sheet_by_index(0)
    sheet.cell_value(0, 0)

    flag = 0

    def check_name (flag):
        name = input('enter user name-english letters only: ')
        for i in range(1, sheet.ncols+1):
            check = sheet.cell_value(i, 0)
            if check == name:
                flag=1
                Password = int(input('Enter a 6-digit password-'))
                index= i
                for j in range(2):
                    if Password == (sheet.cell_value(index, 1)):
                        worker_access = sheet.cell_value(index, 2)
                        Open_Menu(worker_access)
                        break
                    else:
                        Password = int(input('wrong password, try again'))

                if j == 2:
                    print("sorry, too many tries")
                    Error_page()
                break
            else:
                continue

        return flag

    ans=check_name(0)
    if ans==0:
        for i in range (2):
            print("Name does not exist on the system, Try again")
            if check_name(0)!= 0:
                break
Log_In()






'''
find a custumer in the members club
'''
def find_custumer(name, last):
    file_loc = r'C:\Users\micha\Desktop\קוד מיכל\Group2_Yesodot\Hackathon\membership.xlsx'
    workbook = xlrd.open_workbook(file_loc)
    worksheet = workbook.sheet_by_index(0)



